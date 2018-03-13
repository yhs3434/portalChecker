from openpyxl import Workbook
import openpyxl
import datetime
import numpy as np
from url import getURL

class portalChecker:
    def __init__(self,driver,id,pw):
        self.driver=driver
        self.driver.get('https://portal.dankook.ac.kr/web/portal')
        self.driver.implicitly_wait(10)
        print(driver.get_window_size)
        idElement=self.driver.find_element_by_xpath('//*[@id="username"]')
        pwElement=self.driver.find_element_by_xpath('//*[@id="password"]')
        idElement.send_keys(id)
        pwElement.send_keys(pw)
        pwElement.submit()

    def getHouseInfo(self):
        self.driver.get('https://portal.dankook.ac.kr/web/portal/-13')
        self.driver.implicitly_wait(5)
    
        elements=self.driver.find_elements_by_xpath('//*[@class="table_date"]')
        str=""
        for e in elements:
            str=str+e.text+"\n"
        print(str)

    def getHouseInfoC(self):
        self.driver.get('https://portal.dankook.ac.kr/web/portal/-20')
        self.driver.implicitly_wait(5)

        elements=self.driver.find_elements_by_xpath('//*[@id="p_p_id_Bbs_WAR_bbsportlet_"]/div/div/div/div/div[2]/ul/li')
        str=""
        for e in elements:
            m=e.find_element_by_xpath('.//div[2]/a/div[2]')
            str=str+m.text+"\n"
        print(str)

    def checkIsOrNot(self,beforeDate):
        excel_doc=openpyxl.load_workbook('portalCheck.xlsx')
        sheet=excel_doc.active

        now=datetime.datetime.now()
        nowDate=now.strftime('%Y.%m.%d')

        monthDay=now.strftime('%m-%d')
        sheet.cell(row=4,column=4).value=monthDay
        
        timeBefore=datetime.datetime.strptime(beforeDate,'%Y.%m.%d')
        timeCurrent=datetime.datetime.strptime(nowDate,'%Y.%m.%d')
        
        for i in np.arange(47):
            check=False
            self.driver.get(getURL(i))
            self.driver.implicitly_wait(10)
            elements=self.driver.find_elements_by_xpath('//*[@class="table_date"]')
            for e in elements:
                contentDate=e.text
                timeContent=datetime.datetime.strptime(contentDate,'%Y.%m.%d')
                if(timeBefore<=timeContent and timeContent<=timeCurrent):
                    check=True
            
            if(check==True):
                sheet.cell(row=i+5,column=4).value='o'
            else:
                sheet.cell(row=i+5,column=4).value='x'
        
        fileName='portalCheck-'
        fileName+=now.strftime('%m_%d')
        fileName+='.xlsx'
        excel_doc.save(fileName)